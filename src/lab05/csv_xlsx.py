import csv
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Файл не найден: {csv_path}")

    if os.path.getsize(csv_path) == 0:
        raise ValueError("CSV файл пуст")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    with open(csv_path, 'r', encoding='utf-8') as file:
        csv_reader = csv.reader(file)
        rows = list(csv_reader)

    if not rows:
        raise ValueError("CSV файл не содержит данных")

    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max(max_length + 2, 8)
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(xlsx_path)


if __name__ == "__main__":
    try:
        csv_path = Path("data/samples/cities.csv")
        xslx_path = Path("data/out/output.xlsx")
        csv_to_xlsx(csv_path, xslx_path)
        print("Успешно")
    except (ValueError, FileNotFoundError) as e:
        print(f"Ошибка: {e}")